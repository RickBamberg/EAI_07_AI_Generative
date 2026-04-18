"""
agente.py — Núcleo do Agente de Vendas
Lógica extraída do notebook agente_vendas.ipynb.
Usado pelo app.py (Flask) como módulo de backend.
"""

import os
import re
import json
import sqlite3
import io
import base64
import logging
from datetime import datetime

import pandas as pd
import matplotlib
matplotlib.use('Agg')  # backend sem GUI — obrigatório para Flask
import matplotlib.pyplot as plt
from openai import OpenAI

# ─────────────────────────────────────────────────────────────
# CONFIGURAÇÃO
# ─────────────────────────────────────────────────────────────

# Caminhos resolvidos relativos a este arquivo
BASE_DIR  = os.path.dirname(os.path.abspath(__file__))
DB_PATH   = os.path.join(BASE_DIR, 'data', 'DBVendas.db')
LOG_DIR   = os.path.join(BASE_DIR, 'logs')
os.makedirs(LOG_DIR, exist_ok=True)

# Logger de requisições (registra cada pergunta/resposta)
req_logger = logging.getLogger('requisicoes')
req_logger.setLevel(logging.INFO)
fh = logging.FileHandler(os.path.join(LOG_DIR, 'requisicoes.log'), encoding='utf-8')
fh.setFormatter(logging.Formatter('%(asctime)s | %(message)s', datefmt='%Y-%m-%d %H:%M:%S'))
req_logger.addHandler(fh)

# ─────────────────────────────────────────────────────────────
# CONEXÃO COM O BANCO
# ─────────────────────────────────────────────────────────────

def get_connection():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

# ─────────────────────────────────────────────────────────────
# SCHEMA DDL (para o system prompt)
# ─────────────────────────────────────────────────────────────

SCHEMA_DDL = """
-- BANCO: DBVendas.db (SQLite)
-- ATENÇÃO: datas em formato YYYY-MM-DD (texto)
-- Use strftime('%Y', Data_Venda) para filtrar por ano

CREATE TABLE vendas (
    ID_Pedido    INTEGER PRIMARY KEY,
    Data_Venda   TEXT,    -- YYYY-MM-DD
    Data_Entrega TEXT,
    ID_Cliente   INTEGER,
    ID_Canal     INTEGER
);
CREATE TABLE itens_vendas (
    ID_Pedido      INTEGER,
    ID_Produto     INTEGER,
    Qtde           INTEGER,
    Valor_Unitario REAL,
    Valor_Total    REAL
);
CREATE TABLE produtos (
    ID_Produto        INTEGER PRIMARY KEY,
    ID_Subcategoria   INTEGER,
    Descricao_Produto TEXT,
    ID_Marca          INTEGER,
    Preco_Unitario    REAL,
    Tributos          REAL,
    Custo             REAL
);
CREATE TABLE subcategorias (ID_Subcategoria INTEGER PRIMARY KEY, ID_Categoria INTEGER, Subcategoria TEXT);
CREATE TABLE categorias    (ID_Categoria INTEGER PRIMARY KEY, Categoria TEXT);
CREATE TABLE marcas        (ID_Marca INTEGER PRIMARY KEY, Marca TEXT);
CREATE TABLE canais        (ID_Canal INTEGER PRIMARY KEY, Descricao_Canal TEXT);
CREATE TABLE clientes (
    ID_Cliente INTEGER PRIMARY KEY, Nome TEXT, Email TEXT,
    Data_Nascimento TEXT, Estado_Civil TEXT, Genero TEXT,
    Educacao TEXT, ID_Cidade INTEGER
);
CREATE TABLE cidades (ID_Cidade INTEGER PRIMARY KEY, Cidade TEXT, UF TEXT);
CREATE TABLE calendario (
    Data TEXT PRIMARY KEY, Dia INTEGER, Num_Dia_semana INTEGER,
    Dia_da_Semana TEXT, Mes INTEGER, Mes_Ano TEXT, Nome_do_Mes TEXT,
    Trimestre INTEGER, Trimestre_Ano TEXT, Ano TEXT,
    Dia_Util TEXT, Feriado_Nacional TEXT,
    Dias_no_Mes INTEGER, Dia_Util_do_Mes INTEGER, Dias_Uteis_no_Mes INTEGER
);
CREATE TABLE previsao_demanda  (ID_Produto INTEGER, Ano INTEGER, Mes INTEGER, Previsao REAL);
CREATE TABLE backtest_multinivel (Nivel TEXT, Ano INTEGER, MAE REAL, "Erro_%" REAL, Qtd INTEGER, Data_Execucao TEXT);
CREATE TABLE modelo_escolhido (
    Nivel TEXT, Modelo TEXT, Vencedor INTEGER,
    MAE REAL, "Erro_%" REAL, RMSE REAL, R2 REAL,
    N_Treino INTEGER, N_Teste INTEGER, Data_Execucao TEXT
);
"""

# ─────────────────────────────────────────────────────────────
# SYSTEM PROMPT
# ─────────────────────────────────────────────────────────────

SYSTEM_PROMPT = f"""
Você é um agente especialista em análise de dados de vendas do banco DBVendas.db (SQLite).
Responda sempre em português brasileiro.

## SUAS CAPACIDADES
1. `sql_query`     — consulta dados do banco e retorna tabela
2. `gerar_grafico` — gera gráfico como imagem e retorna base64
3. `export_excel`  — exporta dados para arquivo .xlsx

## REGRAS GERAIS
- SEMPRE use uma ferramenta para perguntas sobre dados. Nunca invente números.
- Datas: formato YYYY-MM-DD. Use strftime('%Y', Data_Venda) = '2021' para filtrar ano.
- Joins: vendas ↔ itens_vendas via ID_Pedido; itens_vendas ↔ produtos via ID_Produto
- Previsão: tabela previsao_demanda (ID_Produto, Ano, Mes, Previsao)
- Qualidade modelos: backtest_multinivel (Nivel, Ano, MAE, "Erro_%")
- Modelo vencedor: modelo_escolhido (Nivel, Modelo, Vencedor=1)

## REGRA CRÍTICA — UMA ÚNICA QUERY COMPLETA
NUNCA faça queries exploratórias. Inclua TODOS os campos necessários na primeira query.

❌ ERRADO: 2 queries para a mesma pergunta
✅ CORRETO:
  SELECT p.Descricao_Produto,
         SUM(iv.Qtde)                  AS Total_Vendido,
         ROUND(SUM(iv.Valor_Total), 2) AS Receita_Total,
         COUNT(DISTINCT v.ID_Pedido)   AS Pedidos_Unicos
  FROM itens_vendas iv
  JOIN vendas   v ON iv.ID_Pedido  = v.ID_Pedido
  JOIN produtos p ON iv.ID_Produto = p.ID_Produto
  WHERE strftime('%Y', v.Data_Venda) = '2021'
  GROUP BY p.ID_Produto ORDER BY Total_Vendido DESC LIMIT 5;

## REGRA CRÍTICA — ALINHAMENTO DE TABELAS MARKDOWN
A função `sql_query` já retorna tabelas com alinhamento perfeito, onde:
- As barras verticais `|` ficam alinhadas verticalmente
- Cada coluna tem largura suficiente para comportar o maior conteúdo (cabeçalho OU dados)
- O separador usa `:---` para alinhamento à esquerda

VOCÊ DEVE:
1. Preservar a tabela EXATAMENTE como recebeu, sem modificar os espaços
2. NUNCA adicionar ou remover espaços nas barras `|`
3. Manter o formato de markdown com alinhamento vertical
4. Apresentar a tabela, pular uma linha, e então adicionar os insights

**Principais Insights:**
- Insight 1...
- Insight 2...

## REGRA CRÍTICA — USO DE GRÁFICO
- Use DIRETAMENTE `gerar_grafico` — NÃO chame `sql_query` antes.
- Após o gráfico, forneça SEMPRE análise com 3-5 insights.

## REGRA CRÍTICA — MÚLTIPLAS SÉRIES
Para comparar grupos (ex: Internet vs Loja Física):
- Use `coluna_serie` para pivot automático
- Query com EXATAMENTE 3 colunas: (eixo_x, coluna_serie, valor)
- NÃO inclua colunas extras — causam sobreposição de escalas

✅ CORRETO:
  query="SELECT Mes, Descricao_Canal, SUM(Valor_Total) AS Receita FROM ..."
  coluna_serie="Descricao_Canal"

## REGRA — ANÁLISE APÓS GRÁFICO
Após qualquer gráfico, forneça obrigatoriamente:
- 3 a 5 insights principais
- Maior valor, menor valor, tendência geral
- Uma observação de negócio relevante

## SCHEMA DO BANCO
{SCHEMA_DDL}

## EXEMPLOS DE QUERIES
-- Top 5 produtos 2021:
SELECT p.Descricao_Produto, SUM(iv.Qtde) AS Total_Vendido,
       ROUND(SUM(iv.Valor_Total),2) AS Receita_Total,
       COUNT(DISTINCT v.ID_Pedido) AS Pedidos_Unicos
FROM itens_vendas iv
JOIN vendas v ON iv.ID_Pedido=v.ID_Pedido
JOIN produtos p ON iv.ID_Produto=p.ID_Produto
WHERE strftime('%Y',v.Data_Venda)='2021'
GROUP BY p.ID_Produto ORDER BY Total_Vendido DESC LIMIT 5;

-- Erro % por nível:
SELECT Nivel, ROUND(AVG("Erro_%"),2) AS Erro_Medio,
       ROUND(AVG(MAE),2) AS MAE_Medio, COUNT(*) AS N_Backtests
FROM backtest_multinivel GROUP BY Nivel ORDER BY Erro_Medio;

-- Faturamento por categoria (para gráfico):
SELECT cat.Categoria, ROUND(SUM(iv.Valor_Total),2) AS Receita_Total
FROM itens_vendas iv
JOIN produtos p ON iv.ID_Produto=p.ID_Produto
JOIN subcategorias sub ON p.ID_Subcategoria=sub.ID_Subcategoria
JOIN categorias cat ON sub.ID_Categoria=cat.ID_Categoria
GROUP BY cat.ID_Categoria ORDER BY Receita_Total DESC;
"""

# ─────────────────────────────────────────────────────────────
# FERRAMENTAS
# ─────────────────────────────────────────────────────────────

def sql_query(query: str) -> str:
    try:
        q = query.strip()
        if not q.upper().startswith('SELECT'):
            return 'ERRO: Apenas queries SELECT são permitidas.'
        with get_connection() as conn:
            df = pd.read_sql(q, conn)
        if df.empty:
            return 'A query não retornou resultados.'
        total = len(df)
        tabela = df.head(50).to_string(index=False)
        sufixo = f'\n\n[Mostrando 50 de {total} linhas]' if total > 50 else f'\n\n[{total} linha(s)]'
        return tabela + sufixo
    except Exception as e:
        return f'ERRO ao executar query: {e}\n\nQuery:\n{query}'


def gerar_grafico(query: str, tipo: str = 'bar', titulo: str = '',
                  eixo_x: str = '', eixo_y: str = '',
                  coluna_serie: str = '') -> str:
    """
    Gera gráfico e retorna imagem em base64 (data URI).
    O Flask injeta a imagem diretamente no chat.
    """
    try:
        q = query.strip()
        if not q.upper().startswith('SELECT'):
            return 'ERRO: Apenas queries SELECT são permitidas.'
        with get_connection() as conn:
            df = pd.read_sql(q, conn)
        if df.empty:
            return 'A query não retornou dados para plotar.'
        if len(df.columns) < 2:
            return 'ERRO: Query precisa de pelo menos 2 colunas.'

        col_x = df.columns[0]

        # Decidir colunas Y
        if coluna_serie and coluna_serie in df.columns:
            col_val = [c for c in df.columns if c not in [col_x, coluna_serie]][0]
            df = df.pivot(index=col_x, columns=coluna_serie, values=col_val).reset_index()
            col_x = df.columns[0]
            col_y = df.columns[1:].tolist()
        else:
            numeric_cols = df.select_dtypes(include='number').columns.tolist()
            if not numeric_cols:
                return 'ERRO: Nenhuma coluna numérica encontrada.'
            col_y = [numeric_cols[0]]

        # Plotar
        fig, ax = plt.subplots(figsize=(11, 4.5))
        fig.patch.set_facecolor('#1a1f2e')
        ax.set_facecolor('#1a1f2e')

        cores = ['#4f8ef7', '#f7824f', '#4ff79e', '#f74f9e', '#f7d44f']

        if tipo == 'line':
            for i, c in enumerate(col_y):
                ax.plot(df[col_x].astype(str), df[c], marker='o',
                        label=c, color=cores[i % len(cores)], linewidth=2.5)
            if len(col_y) > 1:
                ax.legend(facecolor='#252b3b', labelcolor='white')
        elif tipo == 'area':
            x_pos = range(len(df))
            for i, c in enumerate(col_y):
                ax.fill_between(x_pos, df[c], alpha=0.35, color=cores[i % len(cores)], label=c)
                ax.plot(x_pos, df[c], color=cores[i % len(cores)], linewidth=2)
            ax.set_xticks(x_pos)
            ax.set_xticklabels(df[col_x].astype(str), rotation=45, ha='right')
            if len(col_y) > 1:
                ax.legend(facecolor='#252b3b', labelcolor='white')
        elif tipo == 'barh':
            if len(col_y) == 1:
                ax.barh(df[col_x].astype(str), df[col_y[0]], color=cores[0])
            else:
                df.set_index(col_x)[col_y].plot(kind='barh', ax=ax,
                    color=cores[:len(col_y)])
                ax.legend(facecolor='#252b3b', labelcolor='white')
        else:  # bar
            if len(col_y) == 1:
                ax.bar(df[col_x].astype(str), df[col_y[0]], color=cores[0])
            else:
                df.set_index(col_x)[col_y].plot(kind='bar', ax=ax,
                    color=cores[:len(col_y)])
                ax.legend(facecolor='#252b3b', labelcolor='white')

        ax.set_title(titulo or f'{col_y[0]} por {col_x}',
                     fontsize=13, fontweight='bold', color='white', pad=12)
        ax.set_xlabel(eixo_x or col_x, fontsize=10, color='#aab4c8')
        ax.set_ylabel(eixo_y or col_y[0], fontsize=10, color='#aab4c8')
        ax.tick_params(colors='#aab4c8')
        for spine in ax.spines.values():
            spine.set_edgecolor('#2e3650')
        ax.grid(axis='y', color='#2e3650', linewidth=0.7, alpha=0.7)

        if tipo != 'area':
            plt.xticks(rotation=45, ha='right', color='#aab4c8')
        plt.tight_layout()

        # Converter para base64
        buf = io.BytesIO()
        plt.savefig(buf, format='png', dpi=130, bbox_inches='tight',
                    facecolor=fig.get_facecolor())
        plt.close()
        buf.seek(0)
        img_b64 = base64.b64encode(buf.read()).decode('utf-8')
        return f'__GRAFICO__data:image/png;base64,{img_b64}'

    except Exception as e:
        return f'ERRO ao gerar gráfico: {e}'


def export_excel(query: str, nome_arquivo: str = 'relatorio',
                 incluir_grafico: bool = True) -> str:
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from openpyxl.chart import BarChart, Reference

        q = query.strip()
        if not q.upper().startswith('SELECT'):
            return 'ERRO: Apenas queries SELECT são permitidas.'
        with get_connection() as conn:
            df = pd.read_sql(q, conn)
        if df.empty:
            return 'A query não retornou dados para exportar.'

        out_dir = os.path.join(BASE_DIR, 'outputs')
        os.makedirs(out_dir, exist_ok=True)
        nome_limpo = nome_arquivo.replace(' ', '_').replace('/', '-')
        caminho = os.path.join(out_dir, f'{nome_limpo}.xlsx')

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Dados'

        header_fill = PatternFill('solid', fgColor='1F4E79')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        center = Alignment(horizontal='center', vertical='center')
        thin = Side(style='thin', color='CCCCCC')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col_name)
            cell.fill, cell.font, cell.alignment, cell.border = (
                header_fill, header_font, center, border)

        alt_fill = PatternFill('solid', fgColor='DCE6F1')
        for row_idx, row in enumerate(df.itertuples(index=False), start=2):
            fill = alt_fill if row_idx % 2 == 0 else PatternFill()
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.fill = fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')

        for col_idx, col_name in enumerate(df.columns, start=1):
            max_len = max(len(str(col_name)),
                          df[col_name].astype(str).str.len().max())
            ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 4, 40)

        if incluir_grafico and len(df) > 1:
            numeric_cols = df.select_dtypes(include='number').columns.tolist()
            if numeric_cols:
                ws_graf = wb.create_sheet('Gráfico')
                chart = BarChart()
                chart.type, chart.title = 'col', nome_arquivo
                chart.y_axis.title = numeric_cols[0]
                chart.x_axis.title = df.columns[0]
                chart.style, chart.width, chart.height = 10, 20, 12
                col_num = df.columns.get_loc(numeric_cols[0]) + 1
                chart.add_data(Reference(ws, min_col=col_num,
                                         min_row=1, max_row=len(df)+1),
                               titles_from_data=True)
                chart.set_categories(Reference(ws, min_col=1,
                                               min_row=2, max_row=len(df)+1))
                ws_graf.add_chart(chart, 'B2')

        wb.save(caminho)
        return f'__EXCEL__{caminho}||{len(df)} linhas exportadas para {nome_limpo}.xlsx'

    except ImportError:
        return 'ERRO: openpyxl não instalado. Execute: pip install openpyxl'
    except Exception as e:
        return f'ERRO ao exportar Excel: {e}'


# Mapa nome → função
FUNCOES = {
    'sql_query':     sql_query,
    'gerar_grafico': gerar_grafico,
    'export_excel':  export_excel,
}

# ─────────────────────────────────────────────────────────────
# SCHEMAS DAS FERRAMENTAS
# ─────────────────────────────────────────────────────────────

TOOLS = [
    {
        'type': 'function',
        'function': {
            'name': 'sql_query',
            'description': (
                'Executa uma query SELECT no DBVendas.db e retorna resultado como tabela. '
                'Use para qualquer pergunta sobre dados de vendas, produtos, clientes, '
                'previsão de demanda e modelos de ML. Gere SQL válido para SQLite. '
                'IMPORTANTE: inclua TODOS os campos necessários em UMA ÚNICA query.'
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'query': {'type': 'string',
                              'description': 'Query SELECT SQLite completa com todos os campos necessários.'}
                },
                'required': ['query']
            }
        }
    },
    {
        'type': 'function',
        'function': {
            'name': 'gerar_grafico',
            'description': (
                'Executa uma query SELECT e gera gráfico visual. '
                'Use DIRETAMENTE quando o usuário pedir gráfico — NÃO chame sql_query antes. '
                'A query deve ter SOMENTE as colunas necessárias para o gráfico. '
                'Para múltiplas séries, use coluna_serie com query de 3 colunas.'
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'query':        {'type': 'string',
                                     'description': 'Query SELECT. Para série única: 2 colunas (X, valor). Para múltiplas séries: 3 colunas (X, coluna_serie, valor).'},
                    'tipo':         {'type': 'string',
                                     'description': "Tipo: 'bar', 'barh', 'line', 'area'. Padrão: 'bar'",
                                     'enum': ['bar', 'barh', 'line', 'area']},
                    'titulo':       {'type': 'string', 'description': 'Título do gráfico'},
                    'eixo_x':       {'type': 'string', 'description': 'Rótulo eixo X'},
                    'eixo_y':       {'type': 'string', 'description': 'Rótulo eixo Y'},
                    'coluna_serie': {'type': 'string',
                                     'description': 'Coluna de agrupamento para múltiplas séries (ex: Descricao_Canal). Query deve ter 3 colunas: (X, coluna_serie, valor).'}
                },
                'required': ['query']
            }
        }
    },
    {
        'type': 'function',
        'function': {
            'name': 'export_excel',
            'description': (
                'Exporta resultado de uma query SELECT para arquivo Excel formatado (.xlsx). '
                'Use quando o usuário pedir exportar, baixar, salvar em Excel ou gerar relatório.'
            ),
            'parameters': {
                'type': 'object',
                'properties': {
                    'query':           {'type': 'string',  'description': 'Query SELECT com os dados a exportar.'},
                    'nome_arquivo':    {'type': 'string',  'description': 'Nome do arquivo sem extensão.'},
                    'incluir_grafico': {'type': 'boolean', 'description': 'Adicionar aba com gráfico. Padrão: True'}
                },
                'required': ['query']
            }
        }
    }
]

# ─────────────────────────────────────────────────────────────
# LOOP REACT — NÚCLEO DO AGENTE
# ─────────────────────────────────────────────────────────────

def _parse_dsml(content: str):
    """Fallback para tool calls em formato DSML do DeepSeek."""
    pattern = r'<\|DSML\|function_calls>.*?<invoke name="([^"]+)">(.*?)</invoke>'
    matches = re.findall(pattern, content, re.DOTALL)
    if not matches:
        return None
    result = []
    for name, body in matches:
        args = {}
        for pname, pval in re.findall(
                r'<parameter name="([^"]+)">(.*?)</parameter>', body, re.DOTALL):
            try:
                args[pname] = json.loads(pval.strip())
            except Exception:
                args[pname] = pval.strip()
        result.append({'name': name, 'arguments': args})
    return result


def _executar_tool(nome: str, args: dict) -> str:
    if nome not in FUNCOES:
        return f'ERRO: ferramenta "{nome}" não reconhecida.'
    try:
        return FUNCOES[nome](**args)
    except Exception as e:
        return f'ERRO ao executar {nome}: {e}'


def responder(pergunta: str, historico: list, llm: OpenAI,
              modelo: str, max_iter: int = 8) -> dict:
    """
    Processa uma pergunta e retorna dict com:
      texto     : resposta em markdown
      grafico   : data URI da imagem (ou None)
      excel     : caminho do arquivo Excel (ou None)
      iteracoes : número de chamadas LLM
      tools_usadas : lista de ferramentas chamadas
    """
    historico = historico + [{'role': 'user', 'content': pergunta}]
    resultado = {
        'texto': '',
        'grafico': None,
        'excel': None,
        'iteracoes': 0,
        'tools_usadas': [],
    }

    for it in range(max_iter):
        resultado['iteracoes'] += 1
        response = llm.chat.completions.create(
            model=modelo,
            messages=[{'role': 'system', 'content': SYSTEM_PROMPT}] + historico,
            tools=TOOLS,
            tool_choice='auto',
            temperature=0,
        )
        msg = response.choices[0].message
        content = msg.content or ''

        # ── CASO A: tool_calls padrão OpenAI ──
        if msg.tool_calls:
            tool_msgs = []
            for tc in msg.tool_calls:
                nome = tc.function.name
                args = json.loads(tc.function.arguments)
                resultado['tools_usadas'].append(nome)
                tool_result = _executar_tool(nome, args)

                # Capturar gráfico/excel do resultado
                if tool_result.startswith('__GRAFICO__'):
                    resultado['grafico'] = tool_result[len('__GRAFICO__'):]
                    tool_result = 'Gráfico gerado com sucesso.'
                elif tool_result.startswith('__EXCEL__'):
                    partes = tool_result[len('__EXCEL__'):].split('||')
                    resultado['excel'] = partes[0]
                    tool_result = partes[1] if len(partes) > 1 else 'Excel exportado.'

                tool_msgs.append({
                    'role': 'tool',
                    'tool_call_id': tc.id,
                    'content': tool_result
                })

            historico.append({'role': 'assistant', 'content': content,
                              'tool_calls': msg.tool_calls})
            historico.extend(tool_msgs)
            continue

        # ── CASO B: DSML ──
        dsml = _parse_dsml(content)
        if dsml:
            tool_msgs = []
            for tc in dsml:
                nome = tc['name']
                args = tc['arguments']
                resultado['tools_usadas'].append(nome)
                tool_result = _executar_tool(nome, args)

                if tool_result.startswith('__GRAFICO__'):
                    resultado['grafico'] = tool_result[len('__GRAFICO__'):]
                    tool_result = 'Gráfico gerado com sucesso.'
                elif tool_result.startswith('__EXCEL__'):
                    partes = tool_result[len('__EXCEL__'):].split('||')
                    resultado['excel'] = partes[0]
                    tool_result = partes[1] if len(partes) > 1 else 'Excel exportado.'

                tool_msgs.append({
                    'role': 'tool',
                    'tool_call_id': f'dsml_{it}',
                    'content': tool_result
                })
            historico.append({'role': 'assistant', 'content': content})
            historico.extend(tool_msgs)
            continue

        # ── CASO C: resposta final ──
        resultado['texto'] = content
        historico.append({'role': 'assistant', 'content': content})

        # Registrar requisição no log
        req_logger.info(
            'PERGUNTA: %s | RESPOSTA: %s | ITER: %d | TOOLS: %s',
            pergunta[:120],
            content[:120],
            resultado['iteracoes'],
            ', '.join(resultado['tools_usadas']) or 'nenhuma'
        )
        return resultado

    resultado['texto'] = 'AVISO: número máximo de iterações atingido.'
    return resultado
